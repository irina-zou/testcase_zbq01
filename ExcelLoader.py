# _*_ coding :utf-8
# @Author : 'zoubq'
# @Time : 2021/4/29 22:33
# @Function :

from openpyxl import load_workbook

from Constant import TEST_EXCEL_DIR_PATH
from model.TestCase import TestCase
import os

IGNORE_SHEETS = ['测试概要', '用例使用说明']


class ExcelLoader(object):

    def __init__(self):
        pass

    def load(self, file_path):
        test_case_list = []
        wb = load_workbook(file_path)
        sheet_num = len(wb.sheetnames)
        for i in range(0, sheet_num):
            sheet_name = wb.sheetnames[i]
            if sheet_name not in IGNORE_SHEETS:
                sheet = wb[sheet_name]
                case_list = self.read_sheet(sheet)
                test_case_list.extend(case_list)
        return test_case_list

    def read_sheet(self, sheet):
        rows_num = sheet.max_row
        columns_num = sheet.max_column
        test_case_list = []

        key_list = []
        for column in range(1, columns_num + 1):
            key_value = sheet.cell(row=1, column=column).value
            key_list.append(key_value)

        for row in range(2, rows_num + 1):
            row_dict = {}
            for column in range(1, columns_num + 1):
                cell_value = sheet.cell(row=row, column=column).value
                row_dict[key_list[column - 1]] = cell_value

            test_case_list.append(TestCase(row_dict))

        return test_case_list

    def batch_load(self, dir_path):
        test_case_list = []
        for dirs, root, test_case_files in os.walk(dir_path):
            for test_case_file in test_case_files:
                # 判断文件结尾的格式，如果是.xlsx就说明是测试用例文件, 如果是~开头说明是临时文件
                if test_case_file.endswith(".xlsx") and not test_case_file.startswith("~"):
                    test_excel_name = os.path.join(dir_path, test_case_file)
                    print(test_excel_name)
                    excel_testcase_list = self.load(test_excel_name)
                    test_case_list.extend(excel_testcase_list)
                else:
                    print("不是测试用例文件")
        return test_case_list
