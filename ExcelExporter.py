# _*_ coding :utf-8
#@Author : 'zoubq'
# @Time : 2021/5/9 22:16
# @Function :
import openpyxl

from model.ExcelResult import RESULT_KEY_LIST


class ExcelExporter(object):
    def __init__(self):
        pass

    def write_to_excel(self, file_path, excel_result_list):
        wb = openpyxl.Workbook()
        ws = wb.active
        for column in range(len(RESULT_KEY_LIST)):
            # excel中的行和列是从1开始计数的，所以需要+1
            ws.cell(1, column + 1).value = RESULT_KEY_LIST[column]

        for row in range(len(excel_result_list)):
            for column in range(len(RESULT_KEY_LIST)):
                print(RESULT_KEY_LIST[column])
                ws.cell(row+2, column+1).value = excel_result_list[row].parse_to_dict()[RESULT_KEY_LIST[column]]

        # 注意，写入后一定要保存
        wb.save(file_path)
        print("成功写入文件: " + file_path + " !")
