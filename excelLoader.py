# _*_ coding :utf-8
#@Author : 'zoubq'
# @Time : 2021/4/29 22:33
# @Function :

from openpyxl import load_workbook

from model.TestCase import TestCase

IGNORE_SHEETS = ['测试概要', '用例使用说明']

class ExcelLoader(object):

    def __init__(self):
        pass

    def load(self, filePath):
        wb = load_workbook(filePath)
        sheet_num = len(wb.sheetnames)
        for i in range (0, sheet_num):
            sheet_name = wb.sheetnames[i]
            if sheet_name not in IGNORE_SHEETS:
                sheet = wb[sheet_name]
                self.read_sheet(sheet)

    def read_sheet(self, sheet):
        rows_num = sheet.max_row
        columns_num = sheet.max_column
        test_case_list = []

        key_list = []
        for column in range(1, columns_num + 1):
            key_value = sheet.cell(row=1, column=column).value
            key_list.append(key_value)

        for row in range(2, rows_num + 1):
            row_dict = {}
            for column in range(1, columns_num + 1):
                cell_value = sheet.cell(row=row, column=column).value
                row_dict[key_list[column-1]] = cell_value

            test_case_list.append(TestCase(row_dict))

        return test_case_list

